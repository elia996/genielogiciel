import serial
import threading
import xlsxwriter
import time
from openpyxl import load_workbook
from openpyxl import Workbook

#userInput = input("Enter the name of the excel file")
filepath = "D:\gamma09.xlsx"
wb=Workbook()
wb.save(filepath)

ser = serial.Serial('COM5',115200)
ser.flushInput()
ser1 = serial.Serial('COM6',115200)
ser1.flushInput()
ser2 = serial.Serial('COM7',115200)
ser2.flushInput()
#workbook = xlsxwriter.Workbook('D:\hello.xlsx')
#worksheet = workbook.add_worksheet()
ctr = 0
packetLimit=1000
dic = {10:0, 20:0,30:0,40:0,50:0,60:0,70:0,80:0,90:0,100:0,
       150:0,200:0,250:0,300:0,400:0,500:0,600:0,700:0,800:0,900:0,1000:0,
       1500:0,2000:0,3000:0,3500:0}

    
def updateDictionary(data):
    for key, value in dic.items():
        if(data<=key):
            dic[key] +=1
    #dic[data]+=1

def mythreadfunction(mySerial):
    mySerial.flushInput()
    global ctr
    while(ctr<packetLimit):
        try:
            #print(mySerial)
            ser_bytes = mySerial.readline()
            line = ser_bytes.decode(encoding='UTF-8')
            data = line.split(",")
            #print(""+line+": "+mySerial.port)
            print(mySerial.port+" : "+line,flush=True)
            updateDictionary(int(data[0])-1000)
            #print(checkData(int(data[0])-1000))
            #updateDictionary(int(data[0])-1000)
            #dic[checkData(int(data[0])-1000)]+=1
            print(dic)
           # lock.acquire()
            ctr = int(data[0])-1000
          #  lock.release()
            #updateDictionary(checkData(int(data[0])-1000))
            #print(dic[checkData(int(data[0]))])
            #print(line)    
           # data = line.split(",")
           # worksheet.write(row,0,float(data[0]))
           # worksheet.write(row,1,float(data[1]))
            #row +=1      
        except:
            print("Keyboard Interrupt")
            break
    print(mySerial.port+" finished!", flush=True)

columnn=1

while True:
    t1 = threading.Thread(target=mythreadfunction, args=(ser,))
    t2 = threading.Thread(target=mythreadfunction, args=(ser1,))
    t3 = threading.Thread(target=mythreadfunction, args=(ser2,))
    t1.daemon = True
    t2.daemon = True
    t3.daemon = True
    t1.start()
    t2.start()
    t3.start()
    t1.join()
    print("t1 killed!")
    t2.join()
    print("t2 killed!")
    t3.join()
    print("t3 killed!")
    print("done ... this is the final dictionary",t3.isAlive())
    print(dic)
    roww =1
    wb=load_workbook(filepath)
    sheet=wb.active
    for key in sorted(dic.keys()):
##        worksheet.write(row,column,float(key))
##        worksheet.write(row,column+1,float(dic[key]))
        sheet.cell(row=roww, column=columnn).value = float(key)
        sheet.cell(row=roww, column=columnn+1).value = float(dic[key])
        roww+=1
        dic[key]=0
    wb.save(filepath)
    columnn+=3
    print("finished writing to excel!")
    ctr=0
    print("going to sleep")
    time.sleep(10)
    #if(column>=10):
    #    workbook.close()

