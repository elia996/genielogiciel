import serial
import threading
import time
from openpyxl import load_workbook
from openpyxl import Workbook

#userInput = input("Enter the name of the excel file")
filepath = "gamma09.xlsx"
wb=Workbook()
wb.save(filepath)

ser = serial.Serial('/dev/cu.usbserial-AI04VD3C',115200)
ser.flushInput()

#workbook = xlsxwriter.Workbook('D:\hello.xlsx')
#worksheet = workbook.add_worksheet()
ctr = 0
myPacketList = list()
firstServerList = list()
secondServerList = list()
thirdServerList = list()
packetLimit=50


def mythreadfunction(mySerial):
    mySerial.flushInput()
    global ctr
    while(ctr<packetLimit):
        try:
            ser_bytes = mySerial.readline()
            line = ser_bytes.decode(encoding='UTF-8')
            data = line.split(",")
            print(line)
            myPacketList.append(int(data[0]))
            firstServerList.append(float(data[1]))
            secondServerList.append(float(data[2]))
            thirdServerList.append(float(data[3]))
            print(mySerial.port+" : "+line,flush=True)
            ctr = int(data[0])-1000     
        except:
            print("Keyboard Interrupt")
            break
    print(mySerial.port+" finished!", flush=True)

columnn=1

while True:
    t1 = threading.Thread(target=mythreadfunction, args=(ser,))
    t1.daemon = True
    t1.start()
    t1.join()
    print("t1 killed!")
    roww =1
    wb=load_workbook(filepath)
    sheet=wb.active
    while(roww<=packetLimit):
        sheet.cell(row=roww, column=columnn).value = myPacketList[roww-1]
        sheet.cell(row=roww, column=columnn+1).value = firstServerList[roww-1]
        sheet.cell(row=roww, column=columnn+2).value = secondServerList[roww-1]
        sheet.cell(row=roww, column=columnn+3).value = thirdServerList[roww-1]
        roww+=1
    wb.save(filepath)
    columnn+=5
    print("finished writing to excel!")
    ctr=0
    print("going to sleep")
    myPacketList = list()
    firstServerList = list()
    secondServerList = list()
    thirdServerList = list()
    ser.flushInput()
    time.sleep(40)
    #if(column>=10):
    #    workbook.close()

